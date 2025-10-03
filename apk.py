import os
import re
import unicodedata
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import shutil
import csv
from openpyxl import load_workbook
import threading
import time
from tkinter import font

class NFGeneratorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("🗂️ Gerador de Pastas de NF - Versão Premium")
        self.root.geometry("950x750")
        self.root.minsize(850, 650)
        
        # Configurar tema e cores
        self.setup_theme()
        
        # Variáveis
        self.empresas_originais = []
        self.empresas_filtradas = []
        self.search_var = tk.StringVar()
        self.filtro_ativo = False  # Flag para controlar filtro inicial
        
        # Configurar interface primeiro
        self.setup_ui()
        
        # Carregar empresas depois da UI
        self.carregar_empresas_csv()
        
        # Configurar trace do filtro DEPOIS de carregar empresas
        self.search_var.trace('w', self.filtrar_empresas)
        
        # Configurar atalhos de teclado
        self.setup_keyboard_shortcuts()
    
    def setup_theme(self):
        """Configura tema visual da aplicação"""
        # Configurar cores do tema
        self.colors = {
            'primary': '#2E86AB',      # Azul principal
            'secondary': '#A23B72',    # Rosa/roxo
            'success': '#27AE60',      # Verde sucesso
            'warning': '#F39C12',      # Laranja aviso
            'danger': '#E74C3C',       # Vermelho perigo
            'light': '#F8F9FA',        # Cinza claro
            'dark': '#2C3E50',         # Azul escuro
            'white': '#FFFFFF',
            'accent': '#3498DB'        # Azul claro
        }
        
        # Configurar estilo ttk
        style = ttk.Style()
        style.theme_use('clam')
        
        # Configurar estilos personalizados
        style.configure('Title.TLabel', 
                       font=('Segoe UI', 18, 'bold'),
                       foreground=self.colors['primary'],
                       background=self.colors['light'],
                       relief='flat',
                       borderwidth=0)
        
        style.configure('Subtitle.TLabel',
                       font=('Segoe UI', 10, 'bold'),
                       foreground=self.colors['dark'],
                       background=self.colors['light'],
                       relief='flat',
                       borderwidth=0)
        
        style.configure('Primary.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white')
        style.map('Primary.TButton',
                 background=[('active', self.colors['accent']),
                           ('!active', self.colors['primary'])])
        
        style.configure('Success.TButton',
                       font=('Segoe UI', 11, 'bold'),
                       foreground='white')
        style.map('Success.TButton',
                 background=[('active', '#229954'),
                           ('!active', self.colors['success'])])
        
        style.configure('Danger.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       foreground='white')
        style.map('Danger.TButton',
                 background=[('active', '#C0392B'),
                           ('!active', self.colors['danger'])])
        
        # Estilo para barra de progresso
        style.configure('Custom.Horizontal.TProgressbar',
                       background=self.colors['primary'],
                       troughcolor=self.colors['light'],
                       borderwidth=0,
                       lightcolor=self.colors['primary'],
                       darkcolor=self.colors['primary'])
        
        # Configurar cor de fundo da janela
        self.root.configure(bg=self.colors['light'])
    
    def carregar_empresas_csv(self):
        """Carrega a lista de empresas do arquivo CSV com melhor tratamento de erros"""
        try:
            # Tentar diferentes caminhos para o arquivo CSV
            possible_paths = [
                os.path.join(os.path.dirname(__file__), "empresas.csv"),  # Mesma pasta do script
                os.path.join(os.path.dirname(os.path.dirname(__file__)), "empresas.csv"),  # Pasta pai
                "empresas.csv"  # Diretório atual
            ]
            
            csv_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    csv_path = path
                    break
            
            if csv_path is None:
                messagebox.showerror("Erro", f"Arquivo empresas.csv não encontrado!\n\nCaminhos verificados:\n" + 
                                   "\n".join(possible_paths) + "\n\nCrie o arquivo CSV com as empresas.")
                self.empresas_originais = []
                return
            
            empresas = []
            with open(csv_path, 'r', encoding='utf-8') as file:
                reader = csv.reader(file)
                try:
                    next(reader)  # Pula o cabeçalho
                except StopIteration:
                    messagebox.showwarning("Aviso", "Arquivo CSV está vazio ou sem cabeçalho.")
                    self.empresas_originais = []
                    return
                
                for linha_num, row in enumerate(reader, start=2):
                    if row and len(row) > 0 and row[0].strip():
                        empresa = row[0].strip()
                        if empresa not in empresas:  # Evita duplicatas
                            empresas.append(empresa)
                        else:
                            print(f"Empresa duplicada encontrada na linha {linha_num}: {empresa}")
            
            self.empresas_originais = sorted(empresas)  # Ordenado alfabeticamente
            self.empresas_filtradas = self.empresas_originais.copy()
            
            if not self.empresas_originais:
                messagebox.showwarning("Aviso", "Nenhuma empresa válida encontrada no arquivo CSV.")
                self.empresas_filtradas = []
            else:
                # Garantir que empresas_filtradas tenha todas as empresas inicialmente
                self.empresas_filtradas = self.empresas_originais.copy()
                print(f"✓ {len(self.empresas_originais)} empresas carregadas com sucesso")
                
                # Atualizar listbox imediatamente após carregar
                if hasattr(self, 'listbox'):
                    self.atualizar_listbox()
                    
                # Atualizar status label se existir
                if hasattr(self, 'status_label'):
                    self.status_label.config(
                        text=f"✅ Mostrando todas as {len(self.empresas_originais)} empresas",
                        bg=self.colors['success'], fg='white'
                    )
                
        except FileNotFoundError:
            messagebox.showerror("Erro", f"Arquivo empresas.csv não encontrado em:\n{csv_path}")
            self.empresas_originais = []
            self.empresas_filtradas = []
        except PermissionError:
            messagebox.showerror("Erro", "Sem permissão para ler o arquivo empresas.csv")
            self.empresas_originais = []
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar empresas do CSV:\n{str(e)}")
            self.empresas_originais = []
    
    def setup_ui(self):
        """Configura a interface do usuário melhorada"""
        # Frame principal com padding e cor de fundo
        main_frame = ttk.Frame(self.root, padding="15")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configurar grid weights para responsividade
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Header com título e subtítulo
        header_frame = ttk.Frame(main_frame)
        header_frame.grid(row=0, column=0, columnspan=3, pady=(0, 25))
        
        title_label = ttk.Label(header_frame, text=" 🗂️ Gerador de Pastas de NF ", 
                               style='Title.TLabel')
        title_label.pack()
        
        subtitle_label = ttk.Label(header_frame, text=" Organize suas notas fiscais de forma automática e eficiente ", 
                                  style='Subtitle.TLabel')
        subtitle_label.pack(pady=(5, 0))
        
        # Frame de busca com visual melhorado
        search_frame = ttk.LabelFrame(main_frame, text="🔍 Buscar Empresas", padding="10")
        search_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 15))
        search_frame.columnconfigure(1, weight=1)
        
        ttk.Label(search_frame, text="📝 Filtro:", font=('Segoe UI', 10, 'bold')).grid(row=0, column=0, padx=(0, 8))
        
        # Entry com placeholder visual
        self.search_entry = ttk.Entry(search_frame, textvariable=self.search_var, 
                                     font=('Segoe UI', 11), width=40)
        self.search_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), padx=(0, 8))
        
        # Botão limpar com ícone
        clear_search_btn = ttk.Button(search_frame, text="🗑️ Limpar", command=self.limpar_busca)
        clear_search_btn.grid(row=0, column=2)
        
        # Adicionar placeholder
        self.search_entry.insert(0, "Digite o nome da empresa...")
        self.search_entry.bind('<FocusIn>', self.on_entry_focus_in)
        self.search_entry.bind('<FocusOut>', self.on_entry_focus_out)
        self.search_entry.config(foreground='gray')
        
        # Frame de seleção de empresas com visual melhorado
        empresas_frame = ttk.LabelFrame(main_frame, text="📋 Selecionar Empresas", padding="10")
        empresas_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 15))
        empresas_frame.columnconfigure(0, weight=1)
        empresas_frame.rowconfigure(0, weight=1)
        
        # Listbox com scrollbar e cores melhoradas
        listbox_frame = ttk.Frame(empresas_frame)
        listbox_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        listbox_frame.columnconfigure(0, weight=1)
        listbox_frame.rowconfigure(0, weight=1)
        
        self.listbox = tk.Listbox(listbox_frame, selectmode=tk.MULTIPLE, 
                                 font=('Segoe UI', 10), height=16,
                                 bg='white', fg=self.colors['dark'],
                                 selectbackground=self.colors['primary'],
                                 selectforeground='white',
                                 activestyle='none',
                                 relief='flat', bd=1)
        self.listbox.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        scrollbar = ttk.Scrollbar(listbox_frame, orient=tk.VERTICAL, command=self.listbox.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.listbox.configure(yscrollcommand=scrollbar.set)
        
        # Botões de seleção com ícones
        selection_frame = ttk.Frame(empresas_frame)
        selection_frame.grid(row=1, column=0, pady=(10, 0))
        
        ttk.Button(selection_frame, text="✅ Selecionar Todas", 
                  command=self.selecionar_todas, style='Primary.TButton').pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(selection_frame, text="❌ Limpar Seleção", 
                  command=self.limpar_selecao).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(selection_frame, text="🔄 Inverter Seleção", 
                  command=self.inverter_selecao).pack(side=tk.LEFT)
        
        # Frame de controles laterais com visual melhorado
        controls_frame = ttk.Frame(main_frame)
        controls_frame.grid(row=2, column=2, sticky=(tk.N, tk.S), padx=(15, 0))
        
        # Adicionar empresa com ícone
        add_frame = ttk.LabelFrame(controls_frame, text="➕ Adicionar Empresa", padding="10")
        add_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.entry_manual = ttk.Entry(add_frame, width=30, font=('Segoe UI', 10))
        self.entry_manual.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Button(add_frame, text="✅ Adicionar", 
                  command=self.adicionar_empresa, style='Primary.TButton').pack(fill=tk.X)
        
        # Excluir empresa com ícone
        delete_frame = ttk.LabelFrame(controls_frame, text="🗑️ Excluir Empresa", padding="10")
        delete_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(delete_frame, text="Selecione empresa(s) na lista", 
                 font=('Segoe UI', 9), foreground=self.colors['secondary']).pack()
        
        ttk.Button(delete_frame, text="🗑️ Excluir Selecionadas", 
                  command=self.excluir_empresa, style='Danger.TButton').pack(fill=tk.X, pady=(8, 0))
        
        # Número da NF com ícone
        nf_frame = ttk.LabelFrame(controls_frame, text="📄 Números da NF", padding="10")
        nf_frame.pack(fill=tk.X, pady=(0, 15))
        
        ttk.Label(nf_frame, text="(separados por vírgula)", 
                 font=('Segoe UI', 9), foreground=self.colors['secondary']).pack()
        
        self.entry_nf = ttk.Entry(nf_frame, width=30, font=('Segoe UI', 10))
        self.entry_nf.pack(fill=tk.X, pady=(8, 0))
        
        # Botões principais com ícones
        buttons_frame = ttk.Frame(controls_frame)
        buttons_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.btn_gerar = ttk.Button(buttons_frame, text="🚀 Gerar Pastas", 
                                   command=self.gerar_pastas_thread, style='Success.TButton')
        self.btn_gerar.pack(fill=tk.X, pady=(0, 8))
        
        ttk.Button(buttons_frame, text="📊 Importar Planilha", 
                  command=self.importar_planilha).pack(fill=tk.X, pady=(0, 8))
        ttk.Button(buttons_frame, text="🧹 Limpar Sessão", 
                  command=self.limpar_sessao).pack(fill=tk.X, pady=(0, 8))
        
        ttk.Button(buttons_frame, text="🔄 Recarregar CSV", 
                  command=self.recarregar_csv).pack(fill=tk.X, pady=(0, 8))
        
        # Barra de progresso com estilo melhorado
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(15, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress = ttk.Progressbar(progress_frame, mode='determinate', style='Custom.Horizontal.TProgressbar', maximum=100)
        self.progress.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Status com cores
        status_frame = ttk.Frame(main_frame)
        status_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E))
        status_frame.columnconfigure(0, weight=1)
        
        self.status_label = tk.Label(status_frame, text="✅ Pronto para uso", 
                                   font=('Segoe UI', 10), 
                                   bg=self.colors['success'], 
                                   fg='white',
                                   relief=tk.FLAT, 
                                   pady=8)
        self.status_label.grid(row=0, column=0, sticky=(tk.W, tk.E))
        
        # Carregar empresas na listbox - sempre mostrar todas inicialmente
        self.atualizar_listbox()
        
        # Configurar status inicial
        if hasattr(self, 'empresas_originais') and self.empresas_originais:
            self.status_label.config(
                text=f"✅ Mostrando todas as {len(self.empresas_originais)} empresas",
                bg=self.colors['success'], fg='white'
            )
        
        # Ativar filtro após carregar empresas
        self.filtro_ativo = True
    
    def setup_keyboard_shortcuts(self):
        """Configura atalhos de teclado"""
        self.root.bind('<Control-f>', lambda e: self.search_entry.focus())
        self.root.bind('<Control-a>', lambda e: self.selecionar_todas())
        self.root.bind('<Control-d>', lambda e: self.limpar_selecao())
        self.root.bind('<Control-r>', lambda e: self.recarregar_csv())
        self.root.bind('<F5>', lambda e: self.recarregar_csv())
        self.root.bind('<Return>', lambda e: self.gerar_pastas_thread() if self.entry_nf.get() else None)
    
    def filtrar_empresas(self, *args):
        """Filtra empresas baseado no texto de busca"""
        # Evitar filtro automático na inicialização
        if not self.filtro_ativo:
            return
            
        busca = self.search_var.get().lower().strip()
        
        # Ignorar placeholder text
        if busca == "digite o nome da empresa...":
            busca = ""
        
        empresas_base = getattr(self, 'empresas_originais', [])
        
        # Se não há empresas carregadas, não fazer nada
        if not empresas_base:
            return
        
        if not busca:
            self.empresas_filtradas = empresas_base.copy()
            self.status_label.config(
                text=f"✅ Mostrando todas as {len(empresas_base)} empresas",
                bg=self.colors['success'], fg='white'
            )
        else:
            self.empresas_filtradas = [emp for emp in empresas_base 
                                     if busca in emp.lower()]
            if self.empresas_filtradas:
                self.status_label.config(
                    text=f"🔍 Filtro ativo: {len(self.empresas_filtradas)} de {len(empresas_base)} empresas",
                    bg=self.colors['warning'], fg='white'
                )
            else:
                self.status_label.config(
                    text="❌ Nenhuma empresa encontrada com esse filtro",
                    bg=self.colors['danger'], fg='white'
                )
        
        self.atualizar_listbox()
    
    def on_entry_focus_in(self, event):
        """Remove placeholder quando campo ganha foco"""
        if self.search_entry.get() == "Digite o nome da empresa...":
            self.search_entry.delete(0, tk.END)
            self.search_entry.config(foreground='black')
    
    def on_entry_focus_out(self, event):
        """Adiciona placeholder quando campo perde foco e está vazio"""
        if not self.search_entry.get():
            self.search_entry.insert(0, "Digite o nome da empresa...")
            self.search_entry.config(foreground='gray')
    
    def limpar_busca(self):
        """Limpa o campo de busca e mostra todas as empresas"""
        self.search_var.set("")
        self.search_entry.delete(0, tk.END)
        self.search_entry.insert(0, "Digite o nome da empresa...")
        self.search_entry.config(foreground='gray')
        
        # Garantir que todas as empresas sejam mostradas
        empresas_base = getattr(self, 'empresas_originais', [])
        if empresas_base:
            self.empresas_filtradas = empresas_base.copy()
            self.status_label.config(
                text=f"✅ Mostrando todas as {len(empresas_base)} empresas",
                bg=self.colors['success'], fg='white'
            )
            self.atualizar_listbox()
        
        self.search_entry.focus()
    
    def atualizar_listbox(self):
        """Atualiza a listbox com as empresas filtradas"""
        self.listbox.delete(0, tk.END)
        empresas_para_mostrar = getattr(self, 'empresas_filtradas', 
                                      getattr(self, 'empresas_originais', 
                                             getattr(self, 'empresas', [])))
        for empresa in empresas_para_mostrar:
            self.listbox.insert(tk.END, empresa)
    
    def selecionar_todas(self):
        """Seleciona todas as empresas visíveis"""
        self.listbox.select_set(0, tk.END)
        self.status_label.config(text=f"Selecionadas {len(self.empresas_filtradas)} empresas")
    
    def limpar_selecao(self):
        """Limpa a seleção atual"""
        self.listbox.selection_clear(0, tk.END)
        self.status_label.config(text="Seleção limpa")
    
    def inverter_selecao(self):
        """Inverte a seleção atual"""
        for i in range(self.listbox.size()):
            if self.listbox.selection_includes(i):
                self.listbox.selection_clear(i)
            else:
                self.listbox.selection_set(i)
        
        selecionadas = len(self.listbox.curselection())
        self.status_label.config(text=f"Seleção invertida - {selecionadas} empresas selecionadas")
    
    def adicionar_empresa(self):
        """Adiciona uma nova empresa à lista e ao CSV"""
        nova_empresa = self.entry_manual.get().strip().upper()
        
        if not nova_empresa:
            messagebox.showwarning("Aviso", "Digite o nome da empresa")
            return
        
        if nova_empresa in self.empresas_originais:
            messagebox.showwarning("Aviso", "Esta empresa já existe na lista")
            return
        
        try:
            # Adicionar à lista em memória
            self.empresas_originais.append(nova_empresa)
            
            # Salvar no CSV reescrevendo o arquivo completo em ordem alfabética
            self.salvar_csv()
            
            # Atualizar interface
            self.filtrar_empresas()
            self.entry_manual.delete(0, tk.END)
            
            # Atualizar status com cor de sucesso
            self.status_label.config(
                text=f"✓ Empresa '{nova_empresa}' adicionada com sucesso",
                bg=self.colors['success'], fg='white'
            )
            
        except Exception as e:
            # Se houve erro ao salvar, remover da lista
            if nova_empresa in self.empresas_originais:
                self.empresas_originais.remove(nova_empresa)
            messagebox.showerror("Erro", f"Erro ao adicionar empresa:\n{str(e)}")
    
    def salvar_csv(self):
        """Salva a lista atual de empresas no arquivo CSV em ordem alfabética"""
        try:
            # Encontrar o caminho do CSV
            possible_paths = [
                os.path.join(os.path.dirname(__file__), "empresas.csv"),
                os.path.join(os.path.dirname(os.path.dirname(__file__)), "empresas.csv"),
                "empresas.csv"
            ]
            
            csv_path = None
            for path in possible_paths:
                if os.path.exists(path):
                    csv_path = path
                    break
            
            if csv_path is None:
                csv_path = os.path.join(os.path.dirname(__file__), "empresas.csv")
            
            # Reescrever o arquivo CSV completo
            with open(csv_path, 'w', newline='', encoding='utf-8') as file:
                writer = csv.writer(file)
                writer.writerow(['empresa'])  # Cabeçalho
                for empresa in sorted(self.empresas_originais):
                    writer.writerow([empresa])
                    
        except Exception as e:
            raise Exception(f"Erro ao salvar CSV: {str(e)}")
    
    def excluir_empresa(self):
        """Exclui empresas selecionadas da lista e do CSV com confirmação"""
        selecionadas = self.listbox.curselection()
        
        if not selecionadas:
            messagebox.showwarning("Aviso", "Selecione pelo menos uma empresa para excluir")
            return
        
        # Obter nomes das empresas selecionadas
        empresas_para_excluir = []
        for index in selecionadas:
            if index < len(self.empresas_filtradas):
                empresas_para_excluir.append(self.empresas_filtradas[index])
        
        if not empresas_para_excluir:
            return
        
        # Confirmação do usuário
        if len(empresas_para_excluir) == 1:
            mensagem = f"Tem certeza que deseja excluir a empresa:\n\n'{empresas_para_excluir[0]}'?\n\nEsta ação não pode ser desfeita."
        else:
            mensagem = f"Tem certeza que deseja excluir {len(empresas_para_excluir)} empresas selecionadas?\n\nEsta ação não pode ser desfeita."
        
        resposta = messagebox.askyesno(
            "Confirmar Exclusão", 
            mensagem,
            icon='warning'
        )
        
        if not resposta:
            return
        
        try:
            # Remover das listas
            for empresa in empresas_para_excluir:
                if empresa in self.empresas_originais:
                    self.empresas_originais.remove(empresa)
            
            # Salvar CSV atualizado
            self.salvar_csv()
            
            # Atualizar interface
            self.filtrar_empresas()
            
            # Atualizar status
            if len(empresas_para_excluir) == 1:
                self.status_label.config(
                    text=f"✓ Empresa '{empresas_para_excluir[0]}' excluída com sucesso",
                    bg=self.colors['success'], fg='white'
                )
            else:
                self.status_label.config(
                    text=f"✓ {len(empresas_para_excluir)} empresas excluídas com sucesso",
                    bg=self.colors['success'], fg='white'
                )
                
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao excluir empresa(s):\n{str(e)}")
    
    def recarregar_csv(self):
        """Recarrega o arquivo CSV"""
        self.status_label.config(text="Recarregando empresas...")
        self.carregar_empresas_csv()
        self.filtrar_empresas()
        self.status_label.config(
            text="✓ Lista de empresas recarregada",
            bg=self.colors['success'], fg='white'
        )
    
    def encontrar_arquivos_nf(self, nf, empresa, usuario):
        """Encontra arquivos de NF nos diretórios especificados"""
        # Varrer toda a raiz "NOTAS FISCAIS" de forma recursiva
        raiz_notas = (
            f"C:\\Users\\{usuario}\\OneDrive - ANFAVEA\\Arquivos's files - FINANCEIRO\\NOTAS FISCAIS"
        )

        def normalize_text(texto: str) -> str:
            # Remove acentos, converte para maiúsculas e mantém apenas letras/números
            if not isinstance(texto, str):
                texto = str(texto)
            nfkd = unicodedata.normalize('NFKD', texto)
            sem_acento = ''.join([c for c in nfkd if not unicodedata.combining(c)])
            upper = sem_acento.upper()
            return re.sub(r"[^A-Z0-9]", "", upper)

        def extract_digits(texto: str) -> str:
            if not isinstance(texto, str):
                texto = str(texto)
            return re.sub(r"\D", "", texto)

        nf_str = str(nf).strip()
        nf_digits = extract_digits(nf_str)
        empresa_norm = normalize_text(empresa)

        arquivos_encontrados = []
        if os.path.exists(raiz_notas):
            for dirpath, dirnames, filenames in os.walk(raiz_notas):
                for arquivo in filenames:
                    nome_norm = normalize_text(arquivo)
                    file_digits = extract_digits(arquivo)

                    # Regra de correspondência: NF por dígitos (mais robusto) e empresa normalizada
                    nf_match = (nf_digits and nf_digits in file_digits) or (nf_str and nf_str in arquivo)
                    empresa_match = empresa_norm in nome_norm

                    if nf_match and empresa_match:
                        arquivos_encontrados.append(os.path.join(dirpath, arquivo))

        return arquivos_encontrados
    
    def gerar_pastas_thread(self):
        """Executa geração de pastas em thread separada"""
        if not self.listbox.curselection():
            messagebox.showwarning("Aviso", "Selecione pelo menos uma empresa")
            return
        
        if not self.entry_nf.get().strip():
            messagebox.showwarning("Aviso", "Digite o(s) número(s) da NF")
            return
        
        # Desabilitar botão e mostrar progresso
        self.btn_gerar.config(state='disabled')
        self.progress.start()
        self.status_label.config(text="Gerando pastas...")
        
        # Executar em thread
        thread = threading.Thread(target=self.gerar_pastas)
        thread.daemon = True
        thread.start()
    
    def gerar_pastas(self):
        """Gera as pastas para as empresas selecionadas"""
        try:
            empresas_selecionadas = [self.empresas_filtradas[i] for i in self.listbox.curselection()]
            numeros_nf = [nf.strip() for nf in self.entry_nf.get().split(',') if nf.strip()]
            
            if len(numeros_nf) != len(empresas_selecionadas):
                self.root.after(0, lambda: messagebox.showwarning(
                    "Aviso", 
                    f"Número de NFs ({len(numeros_nf)}) deve ser igual ao número de empresas selecionadas ({len(empresas_selecionadas)})"
                ))
                return
            
            usuario = os.getlogin()
            pasta_destino = filedialog.askdirectory(title="Selecione onde criar as pastas")
            
            if not pasta_destino:
                return
            
            pastas_criadas = 0
            arquivos_copiados = 0
            erros = []
            
            for i, (empresa, nf) in enumerate(zip(empresas_selecionadas, numeros_nf)):
                try:
                    # Criar pasta
                    nome_pasta = f"{empresa} - NF {nf}"
                    caminho_pasta = os.path.join(pasta_destino, nome_pasta)
                    
                    if not os.path.exists(caminho_pasta):
                        os.makedirs(caminho_pasta)
                        pastas_criadas += 1
                    
                    # Buscar e copiar arquivos
                    arquivos = self.encontrar_arquivos_nf(nf, empresa, usuario)
                    for arquivo in arquivos:
                        nome_arquivo = os.path.basename(arquivo)
                        destino_arquivo = os.path.join(caminho_pasta, nome_arquivo)
                        shutil.copy2(arquivo, destino_arquivo)
                        arquivos_copiados += 1
                    
                    # Atualizar progresso (barra determinate e status)
                    progresso = (i + 1) / len(empresas_selecionadas) * 100
                    self.root.after(0, self.atualizar_progresso, progresso)
                    
                except Exception as e:
                    erros.append(f"{empresa} - NF {nf}: {str(e)}")
            
            # Mostrar resultado
            resultado = f"✓ Concluído!\n"
            resultado += f"Pastas criadas: {pastas_criadas}\n"
            resultado += f"Arquivos copiados: {arquivos_copiados}"
            
            if erros:
                resultado += f"\n\nErros encontrados:\n" + "\n".join(erros[:5])
                if len(erros) > 5:
                    resultado += f"\n... e mais {len(erros) - 5} erros"
            
            self.root.after(0, lambda: messagebox.showinfo("Resultado", resultado))
            self.root.after(0, self.finalizar_progresso)
            
        except Exception as e:
            self.root.after(0, lambda: messagebox.showerror("Erro", f"Erro durante a geração:\n{str(e)}"))
            self.root.after(0, lambda: self.status_label.config(text="Erro na geração"))
        
        finally:
            # Reabilitar interface
            self.root.after(0, lambda: self.progress.stop())
            self.root.after(0, lambda: self.btn_gerar.config(state='normal'))
    
    def importar_planilha(self):
        """Importa dados sem cabeçalho: NF (A) e fornecedor (B),
        selecionando empresas e preenchendo NFs na ordem atual da lista."""
        try:
            arquivo = filedialog.askopenfilename(
                title="Selecione a planilha",
                filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
            )
            
            if not arquivo:
                return
            
            self.status_label.config(text="Importando planilha (sem cabeçalho)...")
            
            workbook = load_workbook(arquivo)
            sheet = workbook.active
            
            # Ler todas as linhas (sem cabeçalho): A=NF, B=Fornecedor
            dados = []
            for row in sheet.iter_rows(min_row=1, values_only=True):
                nf_val = row[0] if len(row) > 0 else None
                emp_val = row[1] if len(row) > 1 else None
                if nf_val and emp_val:
                    nf_str = str(nf_val).strip()
                    emp_str = str(emp_val).strip().upper()
                    if nf_str and emp_str:
                        dados.append((emp_str, nf_str))

            if not dados:
                messagebox.showwarning("Aviso", "Nenhuma linha válida (NF e fornecedor) encontrada na planilha")
                return

            # Mapear empresa -> nf (último valor encontrado caso haja duplicata)
            mapa = {}
            for emp, nf in dados:
                mapa[emp] = str(nf).strip()

            # Selecionar empresas e preencher NFs na ordem da lista filtrada
            self.limpar_selecao()
            ordered_nfs = []
            selected_count = 0
            for i, empresa in enumerate(self.empresas_filtradas):
                if empresa in mapa:
                    self.listbox.selection_set(i)
                    ordered_nfs.append(mapa[empresa])
                    selected_count += 1

            if not ordered_nfs:
                messagebox.showwarning("Aviso", "Nenhuma empresa da planilha coincide com a lista carregada")
                return

            # Preencher campo de NFs com a ordem dos selecionados
            self.entry_nf.delete(0, tk.END)
            self.entry_nf.insert(0, ",".join(ordered_nfs))

            self.status_label.config(text=f"✓ {selected_count} empresas e NFs importados (sem cabeçalho)")
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao importar planilha:\n{str(e)}")
            self.status_label.config(text="Erro na importação")

    def atualizar_progresso(self, p):
        """Atualiza barra de progresso determinate e texto de status."""
        try:
            self.progress.config(mode='determinate')
            try:
                p_float = float(p)
            except Exception:
                p_float = 0
            self.progress['value'] = p_float
            self.status_label.config(text=f"Processando... {p_float:.0f}%")
        except Exception:
            pass

    def finalizar_progresso(self):
        """Reseta barra de progresso e ajusta status para concluído."""
        try:
            self.progress['value'] = 0
            self.status_label.config(text="Concluído com sucesso")
        except Exception:
            pass

    def limpar_sessao(self):
        """Zera todos os campos e estados da sessão atual"""
        try:
            # Limpar busca e restaurar lista
            self.limpar_busca()
            # Limpar seleção de empresas
            self.limpar_selecao()
            # Limpar campo de NF
            if hasattr(self, 'entry_nf'):
                self.entry_nf.delete(0, tk.END)
            # Zerar barra de progresso
            if hasattr(self, 'progress'):
                self.progress.stop()
                self.progress['value'] = 0
                self.progress.config(mode='determinate')
            # Atualizar status
            if hasattr(self, 'status_label'):
                self.status_label.config(text="✅ Sessão limpa e pronta")
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao limpar sessão:\n{str(e)}")

def main():
    root = tk.Tk()
    app = NFGeneratorApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()